"""
绿盟 RSAS (Remote Security Assessment System) 报告解析器

支持绿盟 RSAS 生成的 HTML、XML 和 Excel 三种格式的扫描报告。
"""

import json
import logging
import os
import re
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional
from html import unescape

from . import BaseParser

logger = logging.getLogger(__name__)


class NsfocusParser(BaseParser):
    """绿盟 RSAS 报告解析器（支持 HTML、XML 和 Excel 格式）"""

    # 绿盟风险等级到标准严重程度的映射
    SEVERITY_MAP = {
        '紧急': 'critical',
        '严重': 'critical',
        '高危': 'high',
        '高': 'high',
        '中危': 'medium',
        '中': 'medium',
        '低危': 'low',
        '低': 'low',
        '信息': 'info',
        '提示': 'info',
        'critical': 'critical',
        'high': 'high',
        'medium': 'medium',
        'low': 'low',
        'info': 'info',
        1: 'low',
        2: 'medium',
        3: 'high',
        4: 'critical',
        '1': 'low',
        '2': 'medium',
        '3': 'high',
        '4': 'critical',
    }

    def parse(self, file_path: str) -> List[Dict]:
        """
        解析绿盟 RSAS 报告文件，自动根据扩展名选择解析方式。

        Args:
            file_path: 报告文件路径

        Returns:
            漏洞字典列表
        """
        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext in ('.html', '.htm'):
                return self._parse_html(file_path)
            elif ext == '.xml':
                return self._parse_xml(file_path)
            elif ext in ('.xlsx', '.xls'):
                return self._parse_excel(file_path)
            else:
                # 尝试检测格式
                content = self._read_file(file_path)
                if content:
                    if '<html' in content.lower() or '<!doctype' in content.lower():
                        return self._parse_html(file_path)
                    elif content.strip().startswith('<'):
                        return self._parse_xml(file_path)
                # 尝试 Excel
                return self._parse_excel(file_path)
        except Exception as e:
            logger.error(f"解析绿盟 RSAS 报告文件失败: {e}")
            return []

    def _parse_html(self, file_path: str) -> List[Dict]:
        """
        解析绿盟 RSAS HTML 格式报告。

        使用 BeautifulSoup 解析 HTML 表格，提取漏洞信息。

        Args:
            file_path: HTML 文件路径

        Returns:
            漏洞字典列表
        """
        vulnerabilities = []

        try:
            from bs4 import BeautifulSoup
        except ImportError:
            logger.error("解析 HTML 需要 BeautifulSoup 库，请执行: pip install beautifulsoup4")
            return []

        try:
            content = self._read_file(file_path)
            if not content:
                logger.error(f"无法读取绿盟 RSAS HTML 文件: {file_path}")
                return []

            soup = BeautifulSoup(content, 'html.parser')

            # 查找所有表格
            tables = soup.find_all('table')
            if not tables:
                logger.warning("绿盟 RSAS HTML 中未找到表格")
                return []

            for table in tables:
                rows = table.find_all('tr')
                if not rows:
                    continue

                # 解析表头
                header_row = rows[0]
                headers = []
                for th in header_row.find_all(['th', 'td']):
                    text = th.get_text(strip=True)
                    headers.append(text)

                if not headers:
                    continue

                # 检查是否是漏洞表格（通过表头关键词判断）
                header_text = ' '.join(headers).lower()
                vuln_keywords = ['漏洞', '名称', '等级', '风险', 'cve', '主机', '端口', '描述', '修复']
                match_count = sum(1 for kw in vuln_keywords if kw in header_text)
                if match_count < 2:
                    continue

                # 标准化列名映射
                col_map = self._build_html_col_map(headers)

                # 解析数据行
                for row in rows[1:]:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) < 2:
                        continue

                    row_data = {}
                    for idx, cell in enumerate(cells):
                        if idx < len(headers):
                            col_name = headers[idx]
                            # 获取单元格文本，处理嵌套标签
                            text = cell.get_text(separator=' ', strip=True)
                            row_data[col_name] = text

                    vuln = self._parse_table_row(row_data, col_map)
                    if vuln:
                        vulnerabilities.append(vuln)

        except Exception as e:
            logger.error(f"解析绿盟 RSAS HTML 表格时发生错误: {e}")
            return []

        logger.info(f"绿盟 RSAS HTML 解析完成，共提取 {len(vulnerabilities)} 条漏洞记录")
        return vulnerabilities

    def _build_html_col_map(self, headers: list) -> dict:
        """
        根据表头构建列名映射。

        Args:
            headers: 表头列表

        Returns:
            列名映射字典
        """
        col_map = {}
        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()
            if any(kw in header_lower for kw in ['漏洞名称', '漏洞名', '名称', '漏洞']):
                col_map['name'] = header
            elif any(kw in header_lower for kw in ['风险等级', '等级', '危险等级', '风险', '级别']):
                col_map['severity'] = header
            elif 'cve' in header_lower:
                col_map['cve'] = header
            elif any(kw in header_lower for kw in ['影响主机', '主机', 'ip', '目标', '资产']):
                col_map['host'] = header
            elif any(kw in header_lower for kw in ['端口', 'port']):
                col_map['port'] = header
            elif any(kw in header_lower for kw in ['描述', '详情', '说明', '简介']):
                col_map['description'] = header
            elif any(kw in header_lower for kw in ['修复', '建议', '解决方案', '修补', '处置']):
                col_map['solution'] = header
            elif any(kw in header_lower for kw in ['编号', 'id', '序号']):
                col_map['id'] = header
            elif any(kw in header_lower for kw in ['url', '链接', '地址']):
                col_map['url'] = header
            elif any(kw in header_lower for kw in ['协议', 'protocol']):
                col_map['protocol'] = header
            elif any(kw in header_lower for kw in ['cnvd', 'CNVD']):
                col_map['cnvd'] = header
        return col_map

    def _parse_table_row(self, row_data: dict, col_map: dict) -> Optional[Dict]:
        """
        解析单行表格数据。

        Args:
            row_data: 行数据字典
            col_map: 列名映射

        Returns:
            漏洞字典，数据无效时返回 None
        """
        def get_val(key):
            col = col_map.get(key)
            if col and col in row_data:
                return row_data[col].strip()
            return ''

        name = get_val('name')
        severity_raw = get_val('severity')
        cve = get_val('cve')
        host = get_val('host')
        port = get_val('port')
        description = get_val('description')
        solution = get_val('solution')
        vuln_id = get_val('id')
        url = get_val('url')
        protocol = get_val('protocol')
        cnvd = get_val('cnvd')

        if not name:
            return None

        severity = self.SEVERITY_MAP.get(severity_raw, self._normalize_severity(severity_raw))

        # 处理 CVE
        cve_list = []
        if cve:
            for c in re.findall(r'(CVE-\d{4}-\d+)', cve, re.IGNORECASE):
                cve_list.append(c.upper())

        # 处理 CNVD
        cnvd_list = []
        if cnvd:
            for c in re.findall(r'(CNVD-\d{4}-\d+)', cnvd, re.IGNORECASE):
                cnvd_list.append(c.upper())

        # 构建标题
        title = name
        if vuln_id:
            title = f"[{vuln_id}] {name}"

        return {
            'title': title,
            'severity': severity,
            'host': host,
            'port': port,
            'protocol': protocol,
            'url': url,
            'description': description,
            'solution': solution,
            'cve': ', '.join(sorted(set(cve_list))) if cve_list else '',
            'extra': {
                'scanner': 'nsfocus',
                'vuln_id': vuln_id,
                'cnvd': ', '.join(sorted(set(cnvd_list))) if cnvd_list else '',
                'raw_severity': severity_raw,
            },
        }

    def _parse_xml(self, file_path: str) -> List[Dict]:
        """
        解析绿盟 RSAS XML 格式报告。

        Args:
            file_path: XML 文件路径

        Returns:
            漏洞字典列表
        """
        vulnerabilities = []
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
        except ET.ParseError as e:
            logger.error(f"解析绿盟 RSAS XML 文件失败 (XML格式错误): {e}")
            return []
        except FileNotFoundError:
            logger.error(f"绿盟 RSAS XML 文件不存在: {file_path}")
            return []
        except Exception as e:
            logger.error(f"解析绿盟 RSAS XML 文件时发生未知错误: {e}")
            return []

        try:
            # 尝试多种可能的节点路径
            vuln_nodes = []
            possible_paths = [
                './/vulnerability',
                './/Vulnerability',
                './/vuln',
                './/Vuln',
                './/item',
                './/Item',
                './/result',
                './/Result',
                './/漏洞',
                './/entry',
                './/Entry',
            ]
            for path in possible_paths:
                vuln_nodes = root.findall(path)
                if vuln_nodes:
                    break

            if not vuln_nodes:
                logger.warning("绿盟 RSAS XML 中未找到漏洞节点")
                return []

            for node in vuln_nodes:
                vuln = self._parse_xml_node(node)
                if vuln:
                    vulnerabilities.append(vuln)

        except Exception as e:
            logger.error(f"解析绿盟 RSAS XML 漏洞节点时发生错误: {e}")
            return []

        logger.info(f"绿盟 RSAS XML 解析完成，共提取 {len(vulnerabilities)} 条漏洞记录")
        return vulnerabilities

    def _parse_xml_node(self, node) -> Optional[Dict]:
        """
        解析单个 XML 漏洞节点。

        Args:
            node: XML 漏洞节点

        Returns:
            漏洞字典，解析失败返回 None
        """
        def get_text(tag):
            """获取 XML 标签的文本内容，尝试多种标签名"""
            for t in [tag, tag.lower(), tag.upper(), tag.replace('_', '-')]:
                child = node.find(t)
                if child is not None and child.text:
                    return unescape(child.text.strip())
            return ''

        def get_text_multi(tags):
            """尝试多个标签名"""
            for tag in tags:
                text = get_text(tag)
                if text:
                    return text
            return ''

        name = get_text_multi(['name', 'Name', 'vuln_name', 'vulnName', 'title', 'Title', '漏洞名称'])
        severity_raw = get_text_multi(['severity', 'Severity', 'risk', 'Risk', 'risk_level', 'riskLevel',
                                        'level', 'Level', '风险等级', '危险等级'])
        cve = get_text_multi(['cve', 'CVE', 'cve_id', 'cveId'])
        host = get_text_multi(['host', 'Host', 'ip', 'IP', 'target', 'Target', '影响主机', '主机'])
        port = get_text_multi(['port', 'Port', '端口'])
        description = get_text_multi(['description', 'Description', 'desc', 'Desc', 'detail', 'Detail',
                                       '描述', '详情', '说明'])
        solution = get_text_multi(['solution', 'Solution', 'remediation', 'Remediation', 'fix', 'Fix',
                                    '修复建议', '修复方案', '修补建议'])
        vuln_id = get_text_multi(['id', 'ID', 'vuln_id', 'vulnId', '编号'])
        url = get_text_multi(['url', 'URL', 'link', 'Link', '地址'])
        protocol = get_text_multi(['protocol', 'Protocol', '协议'])

        if not name:
            return None

        severity = self.SEVERITY_MAP.get(severity_raw, self._normalize_severity(severity_raw))

        cve_list = []
        if cve:
            for c in re.findall(r'(CVE-\d{4}-\d+)', cve, re.IGNORECASE):
                cve_list.append(c.upper())

        title = name
        if vuln_id:
            title = f"[{vuln_id}] {name}"

        return {
            'title': title,
            'severity': severity,
            'host': host,
            'port': port,
            'protocol': protocol,
            'url': url,
            'description': description,
            'solution': solution,
            'cve': ', '.join(sorted(set(cve_list))) if cve_list else '',
            'extra': {
                'scanner': 'nsfocus',
                'vuln_id': vuln_id,
                'raw_severity': severity_raw,
            },
        }

    def _parse_excel(self, file_path: str) -> List[Dict]:
        """
        解析绿盟 RSAS Excel 格式报告。

        使用 openpyxl 读取 Excel 文件中的漏洞表格。

        Args:
            file_path: Excel 文件路径

        Returns:
            漏洞字典列表
        """
        vulnerabilities = []

        try:
            import openpyxl
        except ImportError:
            logger.error("解析 Excel 需要 openpyxl 库，请执行: pip install openpyxl")
            return []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"加载绿盟 RSAS Excel 文件失败: {e}")
            return []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # 查找表头行
                header_row_idx = -1
                for idx, row in enumerate(rows):
                    if row:
                        row_text = ' '.join(str(cell) for cell in row if cell).lower()
                        vuln_keywords = ['漏洞', '名称', '等级', '风险', 'cve', '主机', '端口']
                        match_count = sum(1 for kw in vuln_keywords if kw in row_text)
                        if match_count >= 2:
                            header_row_idx = idx
                            break

                if header_row_idx == -1:
                    # 使用第一行作为表头
                    header_row_idx = 0

                headers = []
                header_row = rows[header_row_idx]
                if header_row:
                    headers = [str(cell).strip() if cell else '' for cell in header_row]

                if not headers:
                    continue

                col_map = self._build_html_col_map(headers)

                # 解析数据行
                for row in rows[header_row_idx + 1:]:
                    if not row:
                        continue

                    row_data = {}
                    for idx, cell in enumerate(row):
                        if idx < len(headers):
                            col_name = headers[idx]
                            cell_text = str(cell).strip() if cell is not None else ''
                            row_data[col_name] = cell_text

                    vuln = self._parse_table_row(row_data, col_map)
                    if vuln:
                        vulnerabilities.append(vuln)

        except Exception as e:
            logger.error(f"解析绿盟 RSAS Excel 数据时发生错误: {e}")
            return []
        finally:
            try:
                wb.close()
            except Exception:
                pass

        logger.info(f"绿盟 RSAS Excel 解析完成，共提取 {len(vulnerabilities)} 条漏洞记录")
        return vulnerabilities
