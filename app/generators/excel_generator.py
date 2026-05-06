# -*- coding: utf-8 -*-
"""
Excel 报告生成器 - 使用 openpyxl 生成专业安全测试报告

Excel Report Generator - Generate professional security assessment reports using openpyxl
"""

import os
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

from . import (
    BaseGenerator,
    count_by_level,
    get_risk_bg_color,
)


class ExcelGenerator(BaseGenerator):
    """Excel 格式报告生成器"""

    # 风险等级背景色
    RISK_FILLS = {
        "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "high": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "low": PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid"),
        "info": PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid"),
    }

    # 风险等级字体颜色
    RISK_FONTS = {
        "critical": Font(color="FFFFFF", bold=True, size=10),
        "high": Font(color="FFFFFF", bold=True, size=10),
        "medium": Font(color="333333", bold=True, size=10),
        "low": Font(color="FFFFFF", bold=True, size=10),
        "info": Font(color="FFFFFF", bold=True, size=10),
    }

    # 表头样式
    HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Microsoft YaHei")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 边框
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # 默认单元格样式
    DEFAULT_FONT = Font(size=10, name="Microsoft YaHei")
    DEFAULT_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def generate(
        self,
        project_info: Dict[str, Any],
        vulnerabilities: List[Dict[str, Any]],
        output_path: str,
        lang: str = "zh",
    ) -> str:
        self.lang = lang
        wb = Workbook()

        self._create_vuln_sheet(wb, vulnerabilities)
        self._create_stats_sheet(wb, vulnerabilities)
        self._create_target_sheet(wb, vulnerabilities)

        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        wb.save(output_path)
        return output_path

    def _apply_header_style(self, cell):
        """应用表头样式"""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = self.HEADER_ALIGNMENT
        cell.border = self.THIN_BORDER

    def _apply_cell_style(self, cell, alignment=None):
        """应用默认单元格样式"""
        cell.font = self.DEFAULT_FONT
        cell.alignment = alignment or self.DEFAULT_ALIGNMENT
        cell.border = self.THIN_BORDER

    def _apply_risk_style(self, cell, risk_level: str):
        """应用风险等级样式"""
        level = str(risk_level).lower()
        fill = self.RISK_FILLS.get(level, self.RISK_FILLS["info"])
        font = self.RISK_FONTS.get(level, self.RISK_FONTS["info"])
        cell.fill = fill
        cell.font = font
        cell.alignment = self.DEFAULT_ALIGNMENT
        cell.border = self.THIN_BORDER

    def _create_vuln_sheet(self, wb: Workbook, vulnerabilities: List[Dict[str, Any]]):
        """Sheet 1: 漏洞清单"""
        ws = wb.active
        ws.title = self._t("excel_sheet_vulns")

        # 表头
        headers = [
            self._t("appendix_index"),
            self._t("appendix_cve"),
            self._t("appendix_name"),
            self._t("appendix_level"),
            self._t("appendix_target"),
            self._t("appendix_port"),
            self._t("appendix_protocol"),
            self._t("appendix_desc"),
            self._t("appendix_remediation"),
            self._t("appendix_source"),
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # 数据行
        for row_idx, vuln in enumerate(vulnerabilities, 2):
            risk_level = str(vuln.get("risk_level", "info")).lower()

            values = [
                row_idx - 1,
                vuln.get("cve_id", self._t("na")),
                vuln.get("name", self._t("unknown")),
                self._risk_text(risk_level),
                vuln.get("target", self._t("na")),
                str(vuln.get("port", self._t("na"))),
                vuln.get("protocol", self._t("na")),
                vuln.get("description", self._t("na")),
                vuln.get("remediation", self._t("na")),
                vuln.get("source", self._t("na")),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if col == 4:  # 风险等级列
                    self._apply_risk_style(cell, risk_level)
                elif col in (8, 9):  # 描述和修复建议左对齐
                    self._apply_cell_style(cell, self.LEFT_ALIGNMENT)
                else:
                    self._apply_cell_style(cell)

            # 交替行背景色
            if row_idx % 2 == 0:
                for col in range(1, len(values) + 1):
                    if col != 4:  # 不覆盖风险等级颜色
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                        )

        # 设置列宽
        col_widths = [6, 18, 30, 10, 25, 8, 10, 50, 50, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(vulnerabilities) + 1}"

    def _create_stats_sheet(self, wb: Workbook, vulnerabilities: List[Dict[str, Any]]):
        """Sheet 2: 风险统计"""
        ws = wb.create_sheet(title=self._t("excel_sheet_stats"))

        counts = count_by_level(vulnerabilities)
        total = sum(counts.values())

        # 标题行
        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value=self._t("ch2_title"))
        title_cell.font = Font(size=14, bold=True, color="1A237E", name="Microsoft YaHei")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 表头
        headers = [self._t("ch2_level"), self._t("ch2_count"), self._t("ch2_percentage")]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        # 数据行
        levels = [
            ("critical", self._t("ch2_critical")),
            ("high", self._t("ch2_high")),
            ("medium", self._t("ch2_medium")),
            ("low", self._t("ch2_low")),
            ("info", self._t("ch2_info")),
        ]

        for row_idx, (level_key, level_name) in enumerate(levels, 4):
            count = counts[level_key]
            percentage = f"{count / total * 100:.1f}%" if total > 0 else "0.0%"

            cell_level = ws.cell(row=row_idx, column=1, value=level_name)
            self._apply_risk_style(cell_level, level_key)

            cell_count = ws.cell(row=row_idx, column=2, value=count)
            self._apply_cell_style(cell_count)

            cell_pct = ws.cell(row=row_idx, column=3, value=percentage)
            self._apply_cell_style(cell_pct)

        # 合计行
        total_row = 9
        total_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        total_font = Font(size=11, bold=True, name="Microsoft YaHei")

        cell_total_label = ws.cell(row=total_row, column=1, value=self._t("ch2_total"))
        cell_total_label.fill = total_fill
        cell_total_label.font = total_font
        cell_total_label.alignment = self.DEFAULT_ALIGNMENT
        cell_total_label.border = self.THIN_BORDER

        cell_total_count = ws.cell(row=total_row, column=2, value=total)
        cell_total_count.fill = total_fill
        cell_total_count.font = total_font
        cell_total_count.alignment = self.DEFAULT_ALIGNMENT
        cell_total_count.border = self.THIN_BORDER

        cell_total_pct = ws.cell(row=total_row, column=3, value="100.0%")
        cell_total_pct.fill = total_fill
        cell_total_pct.font = total_font
        cell_total_pct.alignment = self.DEFAULT_ALIGNMENT
        cell_total_pct.border = self.THIN_BORDER

        # 列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

    def _create_target_sheet(self, wb: Workbook, vulnerabilities: List[Dict[str, Any]]):
        """Sheet 3: 按目标统计"""
        ws = wb.create_sheet(title=self._t("excel_sheet_by_target"))

        # 标题行
        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value=self._t("excel_sheet_by_target"))
        title_cell.font = Font(size=14, bold=True, color="1A237E", name="Microsoft YaHei")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 表头
        headers = [
            self._t("appendix_index"),
            self._t("excel_target_col"),
            self._t("excel_vuln_count"),
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        # 统计每个目标的漏洞数量
        target_stats: Dict[str, int] = {}
        target_risk: Dict[str, Dict[str, int]] = {}

        for vuln in vulnerabilities:
            target = vuln.get("target", self._t("unknown"))
            target_stats[target] = target_stats.get(target, 0) + 1

            risk_level = str(vuln.get("risk_level", "info")).lower()
            if target not in target_risk:
                target_risk[target] = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
            if risk_level in target_risk[target]:
                target_risk[target][risk_level] += 1

        # 数据行
        sorted_targets = sorted(target_stats.items(), key=lambda x: x[1], reverse=True)
        for row_idx, (target, count) in enumerate(sorted_targets, 4):
            cell_idx = ws.cell(row=row_idx, column=1, value=row_idx - 3)
            self._apply_cell_style(cell_idx)

            cell_target = ws.cell(row=row_idx, column=2, value=target)
            self._apply_cell_style(cell_target, self.LEFT_ALIGNMENT)

            cell_count = ws.cell(row=row_idx, column=3, value=count)
            self._apply_cell_style(cell_count)

            # 根据最高风险等级设置颜色
            risks = target_risk.get(target, {})
            max_risk = "info"
            for r in ["critical", "high", "medium", "low"]:
                if risks.get(r, 0) > 0:
                    max_risk = r
                    break
            cell_count.fill = PatternFill(
                start_color=get_risk_bg_color(max_risk),
                end_color=get_risk_bg_color(max_risk),
                fill_type="solid",
            )
            if max_risk in ("critical", "high", "low"):
                cell_count.font = Font(color="FFFFFF", bold=True, size=10)
            else:
                cell_count.font = Font(color="333333", bold=True, size=10)

            # 交替行背景色
            if row_idx % 2 == 0:
                cell_idx.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                cell_target.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
